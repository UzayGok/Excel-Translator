
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
import time
import clipboard

error=True
filename=""

while(error):
    try:
        error=False
        filename=input("Enter file name including the extension (ex. translatethis.xlsx): ")
        wb=load_workbook(filename)
    except Exception:
        error=True
        print("File not found, please try again.")
        time.sleep(1)
        

ws=wb.active
driverloc=""
with open("settings.txt") as f:
    driverloc=f.readline()

driver = webdriver.Chrome(driverloc)
driver.get("https://www.deepl.com/translator")
elem = driver.find_element_by_css_selector("div.lmt__inner_textarea_container textarea")

output=driver.find_element_by_id("target-dummydiv")



columns=input("Enter columns(ex. ABDGHI):")
lines=input("Enter lines(ex. 3-36):")
lines=lines.split("-")

assert "DeepL" in driver.title
for column in columns:
    for line in range(int(lines[0]), int(lines[1])+1):
       value= ws[column+str(line)].value
       if(value=="None"):
           continue
       try:
           x=len(ws[column+str(line)].value)
       except TypeError:
           continue
       elem.clear()
       elem.send_keys(value)
       time.sleep(3)
       ws[column+str(line)].value=output.get_attribute('innerHTML')
       
wb.save(filename)